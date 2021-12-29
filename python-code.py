import openpyxl
from PyPDF2 import PdfFileReader, PdfFileWriter

def read_excel(filename,ini_row,ini_col):
    wb = openpyxl.load_workbook(filename)
    sheet_interest = wb.sheetnames[0]
    sheet = wb[sheet_interest]
    max_row = sheet.max_row
    output_list = []
    for i in range(0,max_row-1):
        output_list = output_list + [sheet.cell(row=ini_row+i, column=ini_col).value]
    return output_list


def page_number_start(input_list):
    new_list = [0]
    length = len(input_list)
    for i in range(0,length-1):
        value = new_list[i]+input_list[i]
        new_list = new_list +[value]
    return new_list


def bookmark_pages(master_file,name_list,page_list,output_name):
    length = len(name_list)
    output = PdfFileWriter()  # open output
    input = PdfFileReader(open(master_file, 'rb'))
    n = input.getNumPages()
    for j in range(n):
        output.addPage(input.getPage(j))  # insert page
    for i in range(0,length):
        output.addBookmark(name_list[i], page_list[i], parent=None)
    outputStream = open(output_name, 'wb')  # creating result pdf JCT
    output.write(outputStream)  # writing to result pdf JCT
    outputStream.close()

def add_bookmark_pdf(pdf_file,excel_file,name_row,name_col,page_row,page_col,output):
    namelist = read_excel(excel_file,name_row,name_col)
    pagelist = read_excel(excel_file,page_row,page_col)
    pagelist_final = page_number_start(pagelist)
    bookmark_pages(pdf_file,namelist,pagelist_final,output)





